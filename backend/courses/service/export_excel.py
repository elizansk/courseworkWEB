import datetime
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from django.conf import settings
from django.db.models import Count
from courses.models import Enrollment, Course
import os


def generate_daily_courses_report():
    wb = Workbook()
    wb.remove(wb.active)

    courses = Course.objects.all()

    for course in courses:
        ws = wb.create_sheet(title=f"Course - {course.title[:25]}")
        ws.append([
            "ID", "Email", "Имя", "Фамилия",
            "Статус", "Дата записи", "Прогресс %"
        ])

        enrollments = Enrollment.objects.filter(course=course).select_related('user')

        for e in enrollments:
            ws.append([
                e.id,
                e.user.email,
                e.user.first_name,
                e.user.last_name,
                e.status,
                e.enrolled_at.strftime("%Y-%m-%d") if e.enrolled_at else "",
                e.progress_pct
            ])

    stats_ws = wb.create_sheet(title="Stats")
    stats_ws.append(["Курс", "Количество студентов"])

    stats = (
        Enrollment.objects
        .values('course__title')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')
    )

    for row in stats:
        stats_ws.append([row['course__title'], row['cnt']])

    chart = BarChart()
    chart.title = "Количество студентов по курсам"

    data = Reference(stats_ws, min_col=2, min_row=1, max_row=stats_ws.max_row)
    categories = Reference(stats_ws, min_col=1, min_row=2, max_row=stats_ws.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    stats_ws.add_chart(chart, "D2")

    # сохраняем файл
    reports_dir = os.path.join(settings.BASE_DIR, "reports")
    os.makedirs(reports_dir, exist_ok=True)

    filename = f"courses_report_{datetime.date.today()}.xlsx"
    filepath = os.path.join(reports_dir, filename)

    wb.save(filepath)

    return filepath
